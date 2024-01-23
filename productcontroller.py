from flask import Blueprint, jsonify

from flask import Flask, request, render_template
import os

app = Flask(__name__)


class Employee:
    def __init__(self, empId, empName, empEmail, empGender, fileType=None):
        self.empId = empId
        self.empName = empName
        self.empEmail = empEmail
        self.empGender = empGender

    def __str__(self):
        return f'''(Eid : {self.empId}, ENM : {self.empName}, EMAIL : {self.empEmail}, Gender : {self.empGender})'''

    def __repr__(self):
        return str(self)


e1 = Employee(101, 'AAA', 'abc@gmail.com', 'M')
e2 = Employee(102, 'BBB', 'xxc@gmail.com', 'F')

# FILE_PATH = '/Users/swamirajmathpati/Desktop/flask_proj/Flask_webApp/employee'
# FILE_PATH = os.path.dirname('Flask_webApp')
FILE_PATH = '/usr/src/app/'
JSON_FILE_PATH = FILE_PATH + "employeeemp.json"
TEXT_FILE_PATH = FILE_PATH + "employeeemp.txt"
EXCEL_FILE_PATH = FILE_PATH + "employeeemp.xlsx"
CSV_FILE_PATH = FILE_PATH + "employeeemp.csv"


# JSON_FILE_PATH = FILE_PATH + "emp.json"


def write_into_txt(emp):
    print('Data Writing started in Text file....')
    with open(TEXT_FILE_PATH, 'a') as file:
        empstr = str(emp.empId) + "\t\t" + emp.empName + "\t\t" + emp.empEmail + "\t\t" + emp.empGender + "\n"
        file.writelines(empstr)
    print('Data Writing Completed -- Text')


# write_into_txt(e1)
# import sys
# sys.exit(0)

import json


def write_into_json(emp):
    with open(JSON_FILE_PATH, 'a') as file:
        json.dump(emp.__dict__, file)
        file.writelines("\n")


#
# write_into_json(e2)
# write_into_txt(e1)
# import sys
# sys.exit(0)

def write_into_csv(emp):
    print('Data Writing started in CSV file....')
    with open(CSV_FILE_PATH, 'a') as file:
        empstr = str(emp.empId) + "," + emp.empName + "," + emp.empEmail + "," + emp.empGender + "\n"
        file.writelines(empstr)
    print('Data Writing Completed -- CSV')


# write_into_csv(e2)
# import sys
# sys.exit(0)
import openpyxl


def write_into_excel(emp):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook['emp_data']
        lastrow = sheet.max_row + 1

    except:
        workbook = openpyxl.Workbook()  # BLANK WORKBOOK CREATED
        sheet = workbook.create_sheet('emp_data')  # SHEET IS CREATED
        lastrow = 1

    lastrow = str(lastrow)
    sheet['A' + lastrow] = emp.empId  # CELL MEIN  --. DATA WRITE KR RH HAI
    sheet['B' + lastrow] = emp.empName
    sheet['C' + lastrow] = emp.empEmail
    sheet['D' + lastrow] = emp.empGender
    workbook.save(EXCEL_FILE_PATH)


# write_into_excel(e1)
#
#
# import sys
# sys.exit(0)
def write_into_sqlite3(emp):
    pass


FILE_TYPES_FUN_REF = {
    "J": write_into_json,
    "X": write_into_excel,
    "C": write_into_csv,
    "S": write_into_sqlite3,
    "T": write_into_txt}


@app.route("/")
@app.route("/employee")
@app.route("/employee/save", methods=["GET", "POST"])
def employee_crud_landing_page():
    message = ''
    if request.method == 'POST':
        formdata = request.form
        emp = Employee(**formdata)  # formdata -- dict --> Employee Object
        print('EMPINSTANCE --', emp)
        fileTypes = formdata.getlist('fileType')  # ["J","C","S"]
        if fileTypes:
            for type in fileTypes:
                funref = FILE_TYPES_FUN_REF.get(type)
                funref(emp)
            message = f"Data Written Successfully into specified file formats {fileTypes}"
        else:
            message = "You should select File Types...!"
    return render_template('employee.html', message=message)


def read_json_from_file():
    empList = []
    with open(JSON_FILE_PATH, 'r') as file:
        alllines = file.readlines()

        for line in alllines:
            emp = Employee(**json.loads(line))
            print(emp)
            empList.append(emp)
    return empList


read_json_from_file()


# import sys
# sys.exit(0)

@app.route("/file/read", methods=['POST'])
def papulate_emplist():
    emplist = []
    if request.method == 'POST':
        fileType = request.form.get('fileType')
        if fileType == "J":
            emplist = read_json_from_file()
    return render_template('employee.html', emplist=emplist)


'''

def employee_crud_landing_page():
    message = ''
    if request.method == 'POST':
        formdata = request.form
        Employee(id = formdata.get('eid'))
'''
